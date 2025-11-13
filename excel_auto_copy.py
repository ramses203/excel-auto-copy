#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Auto Copy Tool
1번 엑셀의 각 행 데이터를 2번 엑셀의 1~23행에 반복해서 붙여넣기
"""

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os


class ExcelAutoCopyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Auto Copy Tool")
        self.root.geometry("600x520")
        self.root.resizable(False, False)
        
        self.source_file = ""
        self.target_file = ""
        self.is_running = False
        
        self.setup_ui()
    
    def setup_ui(self):
        # 메인 프레임
        main_frame = tk.Frame(self.root, padx=20, pady=20, bg='white')
        main_frame.pack(fill=tk.BOTH, expand=True)
        self.root.configure(bg='white')
        
        # 제목
        title_label = tk.Label(
            main_frame, 
            text="Excel Auto Copy Tool", 
            font=("Arial", 16, "bold"),
            bg='white',
            fg='black'
        )
        title_label.pack(pady=(0, 20))
        
        # 설명
        desc_text = (
            "Source Excel의 지정한 행 범위를 Target Excel에 반복 복사합니다.\n"
            "예: Source 1~727행 → Target 1행부터 시작 × 23회 = 16,721행"
        )
        desc_label = tk.Label(
            main_frame, 
            text=desc_text, 
            font=("Arial", 9),
            justify=tk.LEFT,
            bg='white',
            fg='black'
        )
        desc_label.pack(pady=(0, 15))
        
        # Source 파일 선택
        source_frame = tk.Frame(main_frame, bg='white')
        source_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(source_frame, text="Source Excel:", width=12, anchor="w", bg='white', fg='black').pack(side=tk.LEFT)
        self.source_label = tk.Label(
            source_frame, 
            text="파일을 선택하세요", 
            bg="#f0f0f0",
            fg='black',
            anchor="w",
            relief=tk.SUNKEN,
            width=40
        )
        self.source_label.pack(side=tk.LEFT, padx=5)
        tk.Button(
            source_frame, 
            text="Browse", 
            command=self.select_source_file
        ).pack(side=tk.LEFT)
        
        # Target 파일 선택
        target_frame = tk.Frame(main_frame, bg='white')
        target_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(target_frame, text="Target Excel:", width=12, anchor="w", bg='white', fg='black').pack(side=tk.LEFT)
        self.target_label = tk.Label(
            target_frame, 
            text="파일을 선택하세요", 
            bg="#f0f0f0",
            fg='black',
            anchor="w",
            relief=tk.SUNKEN,
            width=40
        )
        self.target_label.pack(side=tk.LEFT, padx=5)
        tk.Button(
            target_frame, 
            text="Browse", 
            command=self.select_target_file
        ).pack(side=tk.LEFT)
        
        # 옵션 프레임
        option_frame = tk.LabelFrame(main_frame, text="Options", padx=10, pady=10, bg='white', fg='black')
        option_frame.pack(fill=tk.X, pady=20)
        
        # Source 행 범위 설정
        source_range_frame = tk.Frame(option_frame, bg='white')
        source_range_frame.pack(fill=tk.X, pady=5)
        tk.Label(source_range_frame, text="Source Rows:", width=20, anchor="w", bg='white', fg='black').pack(side=tk.LEFT)
        self.source_start_var = tk.IntVar(value=1)
        tk.Spinbox(
            source_range_frame, 
            from_=1, 
            to=100000, 
            textvariable=self.source_start_var,
            width=8
        ).pack(side=tk.LEFT)
        tk.Label(source_range_frame, text="~", padx=5, bg='white', fg='black').pack(side=tk.LEFT)
        self.source_end_var = tk.IntVar(value=727)
        tk.Spinbox(
            source_range_frame, 
            from_=1, 
            to=100000, 
            textvariable=self.source_end_var,
            width=8
        ).pack(side=tk.LEFT)
        
        # Target 시작 행 설정
        target_start_frame = tk.Frame(option_frame, bg='white')
        target_start_frame.pack(fill=tk.X, pady=5)
        tk.Label(target_start_frame, text="Target Start Row:", width=20, anchor="w", bg='white', fg='black').pack(side=tk.LEFT)
        self.target_start_var = tk.IntVar(value=1)
        tk.Spinbox(
            target_start_frame, 
            from_=1, 
            to=100000, 
            textvariable=self.target_start_var,
            width=10
        ).pack(side=tk.LEFT)
        
        # 반복 횟수 설정
        repeat_frame = tk.Frame(option_frame, bg='white')
        repeat_frame.pack(fill=tk.X, pady=5)
        tk.Label(repeat_frame, text="Repeat Count (per row):", width=20, anchor="w", bg='white', fg='black').pack(side=tk.LEFT)
        self.repeat_var = tk.IntVar(value=23)
        tk.Spinbox(
            repeat_frame, 
            from_=1, 
            to=1000, 
            textvariable=self.repeat_var,
            width=10
        ).pack(side=tk.LEFT)
        
        # 진행 상황
        progress_frame = tk.Frame(main_frame, bg='white')
        progress_frame.pack(fill=tk.X, pady=10)
        
        self.progress_label = tk.Label(
            progress_frame, 
            text="Ready", 
            font=("Arial", 9),
            bg='white',
            fg='black'
        )
        self.progress_label.pack()
        
        self.progress_bar = ttk.Progressbar(
            progress_frame, 
            mode='determinate', 
            length=500
        )
        self.progress_bar.pack(pady=5)
        
        # 실행 버튼
        button_frame = tk.Frame(main_frame, bg='white')
        button_frame.pack(pady=10)
        
        self.start_button = tk.Button(
            button_frame,
            text="Start Copy",
            command=self.start_copy,
            bg="#4CAF50",
            fg="white",
            font=("Arial", 12, "bold"),
            width=15,
            height=2
        )
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = tk.Button(
            button_frame,
            text="Stop",
            command=self.stop_copy,
            bg="#f44336",
            fg="white",
            font=("Arial", 12, "bold"),
            width=15,
            height=2,
            state=tk.DISABLED
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)
    
    def select_source_file(self):
        filename = filedialog.askopenfilename(
            title="Select Source Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.source_file = filename
            self.source_label.config(text=os.path.basename(filename))
    
    def select_target_file(self):
        filename = filedialog.askopenfilename(
            title="Select Target Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.target_file = filename
            self.target_label.config(text=os.path.basename(filename))
    
    def start_copy(self):
        if not self.source_file or not self.target_file:
            messagebox.showerror("Error", "Please select both source and target files!")
            return
        
        if self.source_file == self.target_file:
            messagebox.showerror("Error", "Source and target files must be different!")
            return
        
        # UI 상태 변경
        self.is_running = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.progress_bar['value'] = 0
        
        # 백그라운드에서 실행
        thread = threading.Thread(target=self.copy_process)
        thread.daemon = True
        thread.start()
    
    def stop_copy(self):
        self.is_running = False
        self.progress_label.config(text="Stopping...")
    
    def copy_process(self):
        try:
            # 엑셀 파일 로드
            self.update_progress("Loading source file...", 0)
            source_wb = openpyxl.load_workbook(self.source_file, data_only=True)
            source_ws = source_wb.active
            
            self.update_progress("Loading target file...", 5)
            target_wb = openpyxl.load_workbook(self.target_file)
            target_ws = target_wb.active
            
            # Source 데이터 읽기
            source_start = self.source_start_var.get()
            source_end = self.source_end_var.get()
            
            # 범위 검증
            if source_start > source_end:
                raise Exception(f"Source start row ({source_start}) cannot be greater than end row ({source_end})!")
            
            source_rows = list(source_ws.iter_rows(min_row=source_start, max_row=source_end))
            total_source_rows = len(source_rows)
            
            if total_source_rows == 0:
                raise Exception("No data found in source file!")
            
            repeat_count = self.repeat_var.get()
            target_start_row = self.target_start_var.get()
            total_operations = total_source_rows * repeat_count
            
            self.update_progress(
                f"Processing rows {source_start}~{source_end} ({total_source_rows} rows) × {repeat_count} times...", 
                10
            )
            
            # 복사 작업
            target_row = target_start_row
            completed = 0
            
            for source_idx, source_row_cells in enumerate(source_rows):
                if not self.is_running:
                    break
                
                # 한 행의 데이터를 repeat_count번 반복해서 붙여넣기
                for repeat_idx in range(repeat_count):
                    if not self.is_running:
                        break
                    
                    # 각 셀 복사
                    for col_idx, source_cell in enumerate(source_row_cells, start=1):
                        target_cell = target_ws.cell(row=target_row, column=col_idx)
                        
                        # 값 복사
                        target_cell.value = source_cell.value
                        
                        # 스타일 복사 (옵션)
                        if source_cell.has_style:
                            target_cell.font = source_cell.font.copy()
                            target_cell.border = source_cell.border.copy()
                            target_cell.fill = source_cell.fill.copy()
                            target_cell.number_format = source_cell.number_format
                            target_cell.alignment = source_cell.alignment.copy()
                    
                    target_row += 1
                    completed += 1
                    
                    # 진행률 업데이트 (100번마다)
                    if completed % 100 == 0 or completed == total_operations:
                        progress = 10 + (completed / total_operations * 85)
                        self.update_progress(
                            f"Copying... {completed}/{total_operations} rows ({progress-10:.1f}%)",
                            progress
                        )
            
            if not self.is_running:
                self.update_progress("Stopped by user", 0)
                messagebox.showwarning("Stopped", "Operation stopped by user.")
                return
            
            # 파일 저장
            self.update_progress("Saving file...", 95)
            target_wb.save(self.target_file)
            
            self.update_progress(
                f"Completed! Total {completed} rows copied.", 
                100
            )
            messagebox.showinfo(
                "Success", 
                f"Copy completed successfully!\n\n"
                f"Source: Rows {source_start}~{source_end} ({total_source_rows} rows)\n"
                f"Target: Starting from row {target_start_row}\n"
                f"Total rows copied: {completed}\n"
                f"Target file: {os.path.basename(self.target_file)}"
            )
            
        except Exception as e:
            self.update_progress(f"Error: {str(e)}", 0)
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")
        
        finally:
            # UI 상태 복원
            self.root.after(0, self.reset_ui)
    
    def update_progress(self, text, value):
        def update():
            self.progress_label.config(text=text)
            self.progress_bar['value'] = value
        
        self.root.after(0, update)
    
    def reset_ui(self):
        self.is_running = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)


def main():
    root = tk.Tk()
    app = ExcelAutoCopyApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

